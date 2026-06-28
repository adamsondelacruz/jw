#!/usr/bin/env python3
"""Generate a quarterly maintenance folder (Q1-Q4) for a given year.

Reads Maintenance/Templates/Ashburton Maintenance Calendar .xlsx (latest year
sheet with data; currently 2023) to derive the schedule. A cell filled teal
(#29747D) on the calendar means that job card is scheduled for that month.

For the target quarter, this script:
  1. Selects every job card whose calendar marks any of the quarter's 3 months.
  2. Always includes B-18 (Compliances Monthly) and B-19 (Compliances 3-Monthly).
  3. Copies the matching template .docx from Templates/Job Cards/<Area>/ into
     Maintenance/<year>/q<n>/<Area>/, preserving filenames.
  4. Skips cards whose template does not exist (reported in summary).

Usage:
    python3 generate_quarter.py 2026 3
    python3 generate_quarter.py 2026 3 --dry-run
    python3 generate_quarter.py 2026 3 --calendar-sheet 2023

Re-running on an existing quarter is safe: existing files are NOT overwritten
unless --force is passed.
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]  # Maintenance/
CALENDAR = ROOT / "Templates" / "Ashburton Maintenance Calendar .xlsx"
TEMPLATES_DIR = ROOT / "Templates" / "Job Cards"

AREAS = ["Building", "Civil", "Mechanical", "Electrical", "Equipment", "Technical"]
AREA_HEADER_PREFIXES = {a.upper(): a for a in AREAS}
AREA_HEADER_PREFIXES["TECHNICAL"] = "Technical"  # 2023 sheet has typo "TECHNICAL MAINTENACE"

TEAL_FILL = "FF29747D"  # the colour used on the calendar to mark scheduled months
MONTHS = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN",
          "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")
QUARTER_MONTHS = {
    1: {0, 1, 2},
    2: {3, 4, 5},
    3: {6, 7, 8},
    4: {9, 10, 11},
}

# Always-include compliance cards (run every quarter regardless of calendar)
ALWAYS_INCLUDE = [
    ("Building", "B-18 Compliances (Monthly).docx"),
    ("Building", "B-19 Compliances (3 Monthly).docx"),
]


def code_from_name(name: str) -> str:
    """'B-11 Floor Coverings' -> 'B-11'."""
    return name.strip().split(" ", 1)[0].rstrip(",")


def load_schedule(calendar_path: Path, sheet: str | None) -> dict[str, tuple[str, set[int]]]:
    """Return {full_name: (area, {month_indices})} from the calendar xlsx."""
    wb = load_workbook(calendar_path)
    if sheet is None:
        # pick the most-recent numeric sheet that has any teal-filled cells
        candidates = sorted([s for s in wb.sheetnames if s.isdigit()], reverse=True)
        for s in candidates:
            ws = wb[s]
            for row in ws.iter_rows(min_col=8, max_col=19):
                if any(c.fill.fgColor and getattr(c.fill.fgColor, "rgb", None) == TEAL_FILL for c in row):
                    sheet = s
                    break
            if sheet:
                break
        if not sheet:
            raise SystemExit("No populated year sheet found in calendar.")
    print(f"Using calendar sheet: {sheet}")
    ws = wb[sheet]
    sched: dict[str, tuple[str, set[int]]] = {}
    area = None
    for r in range(1, ws.max_row + 1):
        raw = ws.cell(row=r, column=4).value
        if not raw:
            continue
        name = str(raw).strip()
        up = name.upper()
        # Section header rows e.g. "BUILDING MAINTENANCE"
        matched = next((v for k, v in AREA_HEADER_PREFIXES.items() if up.startswith(k)), None)
        if matched and ("MAINTEN" in up):
            area = matched
            continue
        if area is None:
            continue
        months_marked = set()
        for i, col in enumerate(range(8, 20)):  # H..S = JAN..DEC
            cell = ws.cell(row=r, column=col)
            fg = cell.fill.fgColor.rgb if cell.fill.fgColor and cell.fill.fgColor.type == "rgb" else None
            if fg == TEAL_FILL:
                months_marked.add(i)
        if months_marked:
            sched[name] = (area, months_marked)
    return sched


def template_filename(area: str, calendar_name: str) -> str | None:
    """Find the template .docx whose filename starts with the same code."""
    code = code_from_name(calendar_name)
    area_dir = TEMPLATES_DIR / area
    if not area_dir.is_dir():
        return None
    for f in sorted(area_dir.glob("*.docx")):
        if f.name.startswith(f"{code} "):
            return f.name
    return None


def generate(year: int, quarter: int, dry_run: bool, force: bool, sheet: str | None) -> int:
    if quarter not in QUARTER_MONTHS:
        raise SystemExit("Quarter must be 1, 2, 3, or 4")
    schedule = load_schedule(CALENDAR, sheet)
    target_months = QUARTER_MONTHS[quarter]
    target_dir = ROOT / str(year) / f"q{quarter}"

    # Build the list of (area, template_filename) tuples to copy
    plan: list[tuple[str, str]] = []
    skipped: list[tuple[str, str, str]] = []  # (area, calendar_name, reason)

    for name, (area, sched) in schedule.items():
        if not (sched & target_months):
            continue
        tpl = template_filename(area, name)
        if not tpl:
            skipped.append((area, name, "no template file"))
            continue
        plan.append((area, tpl))

    for area, fname in ALWAYS_INCLUDE:
        if (area, fname) not in plan:
            if (TEMPLATES_DIR / area / fname).is_file():
                plan.append((area, fname))
            else:
                skipped.append((area, fname, "always-include template missing"))

    plan.sort()

    print(f"\nTarget folder: {target_dir}")
    print(f"Months: {', '.join(MONTHS[i] for i in sorted(target_months))}")
    print(f"Planned job cards ({len(plan)}):")
    by_area: dict[str, list[str]] = {}
    for a, f in plan:
        by_area.setdefault(a, []).append(f)
    for a in AREAS:
        items = by_area.get(a, [])
        print(f"  {a} ({len(items)}):")
        for f in items:
            print(f"    - {f}")

    if skipped:
        print(f"\nSkipped ({len(skipped)}):")
        for a, n, why in skipped:
            print(f"  {a:11s} {n}  ({why})")

    if dry_run:
        print("\nDry run - no files written.")
        return 0

    copied = existing = 0
    for area, fname in plan:
        src = TEMPLATES_DIR / area / fname
        dst_dir = target_dir / area
        dst_dir.mkdir(parents=True, exist_ok=True)
        dst = dst_dir / fname
        if dst.exists() and not force:
            existing += 1
            continue
        shutil.copy2(src, dst)
        copied += 1

    print(f"\nWrote {copied} file(s). Skipped {existing} already-existing (use --force to overwrite).")
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("year", type=int)
    p.add_argument("quarter", type=int, choices=[1, 2, 3, 4])
    p.add_argument("--dry-run", action="store_true", help="show plan only")
    p.add_argument("--force", action="store_true", help="overwrite files that already exist")
    p.add_argument("--calendar-sheet", default=None,
                   help="explicit calendar sheet name (default: latest year sheet with teal-filled cells)")
    args = p.parse_args()
    return generate(args.year, args.quarter, args.dry_run, args.force, args.calendar_sheet)


if __name__ == "__main__":
    sys.exit(main())
