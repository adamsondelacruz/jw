#!/usr/bin/env python3
"""Generate the confidential coordinator contact page from workbook sheet 2."""

from pathlib import Path
from html import escape
import re
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT.parent / "2026 CONVENTION COMMITTEE.xlsx"
SHEET = "Elder & MS Contact List"


def phone_href(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if digits.startswith("64"):
        return f"+{digits}"
    if digits.startswith("0"):
        return f"+64{digits[1:]}"
    return f"+64{digits}"


ws = load_workbook(WORKBOOK, read_only=True, data_only=True)[SHEET]
groups: dict[str, dict[str, list[tuple[str, str, str]]]] = {}
area = ""
role = ""

for values in ws.iter_rows(values_only=True):
    cells = [str(value).strip() if value is not None else "" for value in values]
    first = cells[0].strip().upper()
    if first and first not in {"ELDER", "MS"}:
        area = first.title()
        groups.setdefault(area, {"Elder": [], "MS": []})
        continue
    if first in {"ELDER", "MS"}:
        role = "Elder" if first == "ELDER" else "MS"
        continue
    name = cells[1] if len(cells) > 1 else ""
    phone = cells[2] if len(cells) > 2 else ""
    email = cells[3] if len(cells) > 3 else ""
    if area and role and name and name.upper() != "NAME":
        groups[area][role].append((name.strip(), phone.strip(), email.strip()))

total = sum(len(entries) for roles in groups.values() for entries in roles.values())
lines = [
    "# Confidential Elder and Ministerial Servant Contact Masterlist",
    "",
    '<div class="confidential-banner">CONFIDENTIAL — For authorised convention coordination only. Contains personal mobile numbers and jwpub.org addresses. Do not post publicly or distribute beyond those who need it.</div>',
    "",
    f"**Source:** [{WORKBOOK.name}](../2026%20CONVENTION%20COMMITTEE.xlsx), worksheet “{SHEET}”  ",
    f"**Contacts:** {total} · **Areas/congregation groupings:** {len(groups)}  ",
    "**Data treatment:** Names, displayed phone values, email addresses, role categories, and groupings are reproduced from the workbook. Phone links are normalised to New Zealand’s +64 dialling format where possible.",
    "",
    '<div class="contact-tools"><label for="contact-search">Search contacts</label><input id="contact-search" type="search" placeholder="Name, congregation/area, Elder/MS, mobile or email" autocomplete="off"><p class="contact-count" id="contact-count" aria-live="polite"></p></div>',
    "",
]

for area, roles in groups.items():
    area_count = sum(len(entries) for entries in roles.values())
    lines.extend([f'<section class="contact-group" data-group="{escape(area.lower())}">', f"<h2>{escape(area)} ({area_count})</h2>", ""])
    for role, entries in roles.items():
        if not entries:
            continue
        lines.extend([f"<h3>{role} ({len(entries)})</h3>", "", '<table class="contact-table"><thead><tr><th>Name</th><th>Mobile</th><th>jwpub email</th></tr></thead><tbody>'])
        for name, phone, email in entries:
            search = " ".join([area, role, name, phone, email]).lower()
            tel = phone_href(phone)
            lines.append(
                f'<tr class="contact-row" data-search="{escape(search, quote=True)}">'
                f'<td>{escape(name)}</td><td><a href="tel:{escape(tel, quote=True)}">{escape(phone)}</a></td>'
                f'<td><a href="mailto:{escape(email, quote=True)}">{escape(email)}</a></td></tr>'
            )
        lines.extend(["</tbody></table>", ""])
    lines.extend(["</section>", ""])

lines.extend([
    "## Corrections and control",
    "",
    "- Correct the source workbook first, then rerun `generate_contact_masterlist.py` and `render_docs.py`.",
    "- Verify details before urgent or sensitive communication; this page reflects the workbook, not live JW Hub data.",
    "- Apply the current convention direction for access, retention, and destruction of contact information.",
    "",
    "<script>",
    "(() => {",
    "  const input = document.getElementById('contact-search');",
    "  const count = document.getElementById('contact-count');",
    "  const rows = [...document.querySelectorAll('.contact-row')];",
    "  const groups = [...document.querySelectorAll('.contact-group')];",
    "  function filter() {",
    "    const terms = input.value.toLowerCase().trim().split(/\\s+/).filter(Boolean);",
    "    let visible = 0;",
    "    for (const row of rows) { const match = terms.every(term => row.dataset.search.includes(term)); row.hidden = !match; if (match) visible++; }",
    "    for (const group of groups) group.hidden = !group.querySelector('.contact-row:not([hidden])');",
    "    count.textContent = `${visible} of ${rows.length} contacts shown`;",
    "  }",
    "  input.addEventListener('input', filter); filter();",
    "})();",
    "</script>",
])

(ROOT / "contact-masterlist.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
print(f"Generated {total} contacts across {len(groups)} groupings.")
